import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Any, List
from app.models.finance_models import FinancialModel

class ExcelService:
    """
    Generates Excel files matching the original format based on the financial model.
    """

    def generate_excel(self, model: FinancialModel) -> bytes:
        """
        Generates an Excel file from the financial model's projections.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Financial Model"
        
        # Create headers
        self._create_headers(ws, len(model.monthly_projections))
        
        # Populate data
        self._populate_data(ws, model)
        
        # Style the worksheet
        self._apply_styling(ws)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def _create_headers(self, ws, num_months: int):
        """
        Creates the header row for the Excel sheet dynamically based on the number of months.
        """
        headers = ["Metric", "Unit"] + [f"M{i}" for i in range(1, num_months + 1)]
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for col_num, header_text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header_text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def _populate_data(self, ws, model: FinancialModel):
        """
        Populates data rows matching the original Excel structure.
        """
        row = 2
        
        # Helper function to write a row of data
        def write_row(metric_name: str, unit: str, values: List[Any]):
            nonlocal row
            ws.cell(row=row, column=1, value=metric_name)
            ws.cell(row=row, column=2, value=unit)
            for col, value in enumerate(values, 3):
                ws.cell(row=row, column=col, value=value)
            row += 1

        # Extract data from the model's monthly projections
        projections = model.monthly_projections
        
        # List of values to be extracted and their corresponding labels
        data_rows = [
            ("# of sales people", "#", [p.sales_people for p in projections]),
            ("# of large customer accounts they can sign per month, sales person", "#", [model.assumptions.get("large_customers_per_salesperson", 1.5) for _ in projections]),
            ("# of large customer accounts onboarded per month", "#", [p.large_customers_acquired for p in projections]),
            ("Cumulative # of large paying customers", "#", [p.large_customers_cumulative for p in projections]),
            ("Average revenue per large customer", "$ per month", [p.large_customer_revenue / p.large_customers_cumulative if p.large_customers_cumulative > 0 else 0 for p in projections]),
            ("Digital Marketing spend per month", "$ per month", [p.marketing_spend for p in projections]),
            ("Average CAC", "$ per customer", [model.assumptions.get("cac", 1500) for _ in projections]),
            ("# of sales inquiries", "#", [model.assumptions.get("sales_inquiries_per_month", 160) for _ in projections]),
            ("% conversions from demo to sign ups", "%", [f"{model.assumptions.get('demo_rate', 0.45) * 100}%" for _ in projections]),
            ("# of small/medium paying customers onboarded", "#", [p.small_customers_acquired for p in projections]),
            ("Cumulative number of small/medium paying customers", "#", [p.small_customers_cumulative for p in projections]),
            ("Average revenue per small/medium customer", "$ per customer", [p.small_customer_revenue / p.small_customers_cumulative if p.small_customers_cumulative > 0 else 0 for p in projections]),
            ("Revenue from large clients", "$ per month", [p.large_customer_revenue for p in projections]),
            ("Revenue from small and medium clients", "$ per month", [p.small_customer_revenue for p in projections]),
            ("Total Revenues", "$ per month", [p.total_revenue for p in projections]),
            ("Total Revenues", "$ Mn per month", [round(p.total_revenue / 1000000, 2) for p in projections])
        ]
        
        for metric, unit, values in data_rows:
            write_row(metric, unit, values)

    def _apply_styling(self, ws):
        """
        Applies basic styling to match the original Excel format.
        """
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column].width = min(adjusted_width, 50)
            
        # Center align numeric data
        for row in ws.iter_rows(min_row=2, min_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")
